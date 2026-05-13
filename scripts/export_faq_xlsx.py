"""Export FAQ JSONs to XLSX for human evaluation. Run from project root:
    python scripts/export_faq_xlsx.py faq_data
    python scripts/export_faq_xlsx.py faq_data --force
    python scripts/export_faq_xlsx.py "faq_data/PPC/PPC - Bacharelado em Ciências de Computação (BCC) 2026-1.json"
"""

import json
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from shared.utils import load_config, to_output_path


# (column_name, width, wrap_text)
COLUMNS: list[tuple[str, int, bool]] = [
    ("chunk_pages", 12, False),
    ("source_page", 12, False),
    ("question", 60, True),
    ("answer", 80, True),
]


def _rows_for_doc(faq_doc: dict) -> list[dict]:
    """Flatten one FAQ JSON dict into a list of row dicts (one per QA pair)."""
    rows: list[dict] = []
    for chunk in faq_doc.get("chunks", []):
        chunk_pages = ", ".join(str(p) for p in chunk.get("pages", []))
        for qa in chunk.get("qa_pairs", []):
            rows.append({
                "chunk_pages": chunk_pages,
                "source_page": qa.get("source_page", ""),
                "question": qa.get("question", ""),
                "answer": qa.get("answer", ""),
            })
    rows.sort(key=lambda r: r["source_page"] if isinstance(r["source_page"], int) else float("inf"))
    return rows


def _write_workbook(rows: list[dict], out_path: Path) -> None:
    """Write rows to an XLSX with header formatting and wrapped text."""
    wb = Workbook()
    ws = wb.active
    ws.title = "QA"

    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment = Alignment(vertical="top")

    for col_idx, (name, width, _wrap) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (name, _width, wrap) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[name])
            cell.alignment = wrap_alignment if wrap else top_alignment

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="File or directory under faq_data/")
    parser.add_argument("--force", action="store_true", help="Overwrite existing xlsx outputs")
    args = parser.parse_args()

    config = load_config()
    faq_eval_dir = Path(config["paths"]["faq_eval_dir"])

    in_path: Path = args.input
    if in_path.is_file():
        files = [in_path] if in_path.suffix.lower() == ".json" else []
    else:
        files = sorted(in_path.rglob("*.json"))

    if not files:
        print(f"No .json files found at {in_path}")
        return

    input_root = Path(in_path.parts[0])

    skipped = 0
    failed: list[str] = []
    empty: list[str] = []
    for file in tqdm(files, desc="Exporting xlsx", unit="file"):
        out_file = to_output_path(file, input_root, faq_eval_dir).with_suffix(".xlsx")
        if out_file.exists() and not args.force:
            tqdm.write(f"  {file.name} -> SKIP (exists; use --force to overwrite)")
            skipped += 1
            continue

        tqdm.write(f"  {file.name} -> {out_file}")
        try:
            with file.open("r", encoding="utf-8") as f:
                faq_doc = json.load(f)
            rows = _rows_for_doc(faq_doc)
            _write_workbook(rows, out_file)
            if not rows:
                empty.append(file.name)
                tqdm.write("    WARNING: no QA pairs in this doc")
        except Exception as e:
            failed.append(file.name)
            tqdm.write(f"    FAILED: {e}")

    processed = len(files) - skipped - len(failed)
    print(f"\nDone. {processed} processed, {skipped} skipped, {len(failed)} failed.")
    if empty:
        print(f"Docs with zero QA pairs: {', '.join(empty)}")
    if failed:
        print(f"Failed docs: {', '.join(failed)}")


if __name__ == "__main__":
    main()
